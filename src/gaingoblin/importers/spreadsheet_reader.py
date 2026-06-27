from __future__ import annotations

import csv
from pathlib import Path

from gaingoblin.importers.import_models import RawImportRow

SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


def read_spreadsheet(path: Path | str) -> list[RawImportRow]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        return read_csv(source)
    if suffix == ".xlsx":
        return read_xlsx(source)
    raise ValueError(f"Unsupported import file type: {suffix}")


def read_csv(path: Path) -> list[RawImportRow]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            return []
        return [
            RawImportRow(
                row_number=index,
                values={key: _stringify(value) for key, value in row.items() if key is not None},
            )
            for index, row in enumerate(reader, start=2)
        ]


def read_xlsx(path: Path) -> list[RawImportRow]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("Excel import requires openpyxl. Install GainGoblin with Excel support.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [_stringify(value) for value in next(rows)]
    except StopIteration:
        workbook.close()
        return []

    results: list[RawImportRow] = []
    for row_number, values in enumerate(rows, start=2):
        row_values = {
            header: _stringify(value)
            for header, value in zip(headers, values)
            if header
        }
        if any(value.strip() for value in row_values.values()):
            results.append(RawImportRow(row_number=row_number, values=row_values))
    workbook.close()
    return results


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
